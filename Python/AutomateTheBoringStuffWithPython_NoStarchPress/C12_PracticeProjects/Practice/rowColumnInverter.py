import openpyxl
wb=openpyxl.load_workbook('produceSales.xlsx')
print(wb.sheetnames)
sheet=wb['Sheet']
inverted=openpyxl.Workbook()
inverted_sheet=inverted.active
print(sheet.max_column,sheet.max_row)
for x in range(1,sheet.max_column+1):
    for y in range(1,150+1):#too big to transpose,limiting to 150
        inverted_sheet.cell(row=x,column=y).value= sheet.cell(row=y,column=x).value

inverted.save('transposed_produced_sales.xlsx')
