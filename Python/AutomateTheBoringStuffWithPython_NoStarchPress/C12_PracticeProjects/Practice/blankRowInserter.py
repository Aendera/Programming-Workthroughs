import openpyxl
import sys

N,M,workbook=sys.argv[1:4]
wb=openpyxl.load_workbook(workbook)
sheet=wb.active

alt=openpyxl.Workbook()
alt_sheet=alt.active
for row in range(1,sheet.max_row):
    for column in range(1,sheet.max_column+1):
        if row<=int(N):
            alt_sheet.cell(row=row,column=column).value = sheet.cell(row=row,column=column).value
        else:
            alt_sheet.cell(row=row+int(M),column=column).value = sheet.cell(row=row,column=column).value
alt.save('extrablanklinesin'+workbook)